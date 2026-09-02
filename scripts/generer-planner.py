#!/usr/bin/env python3
"""
Génère les fichiers nécessaires au suivi de projet dans Microsoft Planner :
  - une feuille de calcul détaillée (Excel)
  - un CSV complet
  - un fichier texte par compartiment (bucket), à coller directement dans Planner

Usage : python3 scripts/generer-planner.py
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

ICI = os.path.dirname(os.path.abspath(__file__))
SORTIE = os.path.join(ICI, "..", "docs", "planner")
os.makedirs(SORTIE, exist_ok=True)

VIOLET = "500070"
VERTF = "16a34a"
AMBREF = "b45309"
GRISF = "475569"

# ------------------------------------------------------------------ données --
# (module, tâche, bucket, avancement %, échéance, priorité, responsable, notes)
TACHES = [
    # ---------------- LIVRÉ ----------------
    ("Cadrage", "Cadrage et analyse des besoins", "Livré", 100, "28/08/2026", "Haute", "Projet", "Objectifs, périmètre, règles de gestion"),
    ("Cadrage", "Choix de l'architecture (Next.js + Supabase)", "Livré", 100, "28/08/2026", "Haute", "DSI", "Application web + base temps réel"),
    ("Sécurité", "Authentification et gestion des 3 rôles", "Livré", 100, "29/08/2026", "Haute", "DSI", "Demandeur / Responsable / Administrateur"),
    ("Demandes", "Module demandes d'achat (assistant 6 étapes)", "Livré", 100, "29/08/2026", "Haute", "DSI", "Création, brouillon, reprise"),
    ("Demandes", "Numérotation automatique ACH-AAAA-NNNNN", "Livré", 100, "29/08/2026", "Haute", "DSI", "Compteur partagé, sans doublon"),
    ("Demandes", "Rendre le coût présumé facultatif", "Livré", 100, "01/09/2026", "Moyenne", "DSI", "Demande créable sans montant"),
    ("Fournisseurs", "Catalogue fournisseurs (création, modification, statut)", "Livré", 100, "30/08/2026", "Haute", "DSI", "Nom, références, emplacement, WhatsApp, site"),
    ("Fournisseurs", "Import des fournisseurs (Excel, CSV, JSON)", "Livré", 100, "30/08/2026", "Haute", "DSI", "Aperçu, lignes invalides signalées"),
    ("Fournisseurs", "Export des fournisseurs et modèles de fichiers", "Livré", 100, "30/08/2026", "Moyenne", "DSI", "Excel, JSON, CSV"),
    ("Négociation", "Grille de négociation article × fournisseur", "Livré", 100, "31/08/2026", "Haute", "DSI", "Un prix par article et par fournisseur"),
    ("Négociation", "Rendre les prix facultatifs (seuil paramétrable)", "Livré", 100, "02/09/2026", "Haute", "DSI", "Seuil 0 = facultatif"),
    ("Négociation", "Coût total rendu (remise, TVA, frais de livraison)", "Livré", 100, "31/08/2026", "Haute", "DSI", "Comparaison fiable des offres"),
    ("Comparaison", "Score multicritère /100 et classement", "Livré", 100, "31/08/2026", "Moyenne", "DSI", "Prix, délai, garantie, paiement"),
    ("Comparaison", "Tableau croisé et meilleurs prix surlignés", "Livré", 100, "31/08/2026", "Haute", "DSI", "Lecture seule, étape non bloquante"),
    ("Validation", "Circuit de validation (approuver / refuser / modifier)", "Livré", 100, "01/09/2026", "Haute", "DSI", "Motif obligatoire si refus"),
    ("Commandes", "Bons de commande BC-AAAA-NNNNN et confirmation", "Livré", 100, "01/09/2026", "Haute", "DSI", "Lignes aux prix négociés"),
    ("Commandes", "Commande et impression sans fournisseur ni prix", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Cas des consultations en cours"),
    ("Réceptions", "Réceptions complètes et partielles", "Livré", 100, "01/09/2026", "Haute", "DSI", "Clôture automatique si complète"),
    ("Administration", "Suppressions (demande, commande, réception, utilisateur)", "Livré", 100, "01/09/2026", "Haute", "DSI", "Cascades et garde-fous"),
    ("Alertes", "Notifications et relances automatiques", "Livré", 100, "02/09/2026", "Haute", "DSI", "Validation et réception, anti-doublon"),
    ("Alertes", "Bandeau d'alertes et pastilles sur le tableau de bord", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Vue immédiate des retards"),
    ("Alertes", "Sons de notification (3 niveaux) et volume réglable", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Avec son de secours si bloqué"),
    ("Alertes", "Alarme répétée jusqu'à lecture du message", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "15 s, maximum 10 répétitions"),
    ("Alertes", "Notifications du navigateur (bulles système)", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Actives onglet en arrière-plan"),
    ("Alertes", "Alerte visuelle (onglet clignotant, pastille pulsante)", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Secours si le son est coupé"),
    ("Documents", "Impression fiche de demande et bon de commande", "Livré", 100, "02/09/2026", "Haute", "DSI", "Colonnes de prix vides à compléter"),
    ("Documents", "Exports PDF (demandes, commandes, réceptions)", "Livré", 100, "01/09/2026", "Haute", "DSI", "jsPDF, repli impression"),
    ("Documents", "Exports Excel et CSV", "Livré", 100, "01/09/2026", "Moyenne", "DSI", "Comparaison incluse"),
    ("Identité", "Intégration du logo et charte violet / bleu nuit", "Livré", 100, "02/09/2026", "Moyenne", "Direction", "Site, impressions, PDF, favicon"),
    ("Données", "Synchronisation Supabase en temps réel", "Livré", 100, "02/09/2026", "Haute", "DSI", "Partage entre tous les postes"),
    ("Données", "Sécurisation anti-écrasement entre appareils", "Livré", 100, "02/09/2026", "Haute", "DSI", "Le serveur fait foi au démarrage"),
    ("Pilotage", "Tableau de bord graphique (3 graphiques)", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "Achats 12 mois, répartition, top 5"),
    ("Qualité", "Suite de tests automatisés (197 contrôles)", "Livré", 100, "02/09/2026", "Haute", "DSI", "13 séries, exécutées avant publication"),
    ("Déploiement", "Déploiement continu GitHub vers Vercel", "Livré", 100, "02/09/2026", "Haute", "DSI", "Une commande : scripts/publish.mjs"),
    ("Documentation", "Cahier des charges et synthèse direction", "Livré", 100, "02/09/2026", "Moyenne", "DSI", "PDF et Word, 14 et 4 pages"),

    # ---------------- EN COURS ----------------
    ("Qualité", "Recette utilisateur et ajustements", "En cours", 60, "05/09/2026", "Haute", "Direction", "Retours des utilisateurs finaux"),
    ("Sécurité", "Durcissement des politiques d'accès (RLS)", "En cours", 30, "09/09/2026", "Haute", "DSI", "Restreindre par rôle avant généralisation"),
    ("Sécurité", "Journal d'audit des actions sensibles", "En cours", 15, "12/09/2026", "Haute", "DSI", "Qui a supprimé/modifié quoi et quand"),

    # ---------------- PLANIFIÉ ----------------
    ("Sécurité", "Authentification sécurisée (mots de passe hachés)", "Planifié", 0, "19/09/2026", "Haute", "DSI", "Migration vers Supabase Auth"),
    ("Alertes", "Notifications par e-mail", "Planifié", 0, "26/09/2026", "Moyenne", "DSI", "Réduction des délais de réponse"),
    ("Pilotage", "Budgets par service avec alerte de dépassement", "Planifié", 0, "10/10/2026", "Moyenne", "Direction", "Pilotage financier"),
    ("Fournisseurs", "Portail fournisseur (saisie directe des offres)", "Planifié", 0, "24/10/2026", "Basse", "DSI", "Productivité sur les consultations"),
    ("Logistique", "Gestion des stocks liée aux réceptions", "Planifié", 0, "31/10/2026", "Basse", "Logistique", "À cadrer"),
    ("Sécurité", "Authentification unique (SSO)", "Planifié", 0, "31/10/2026", "Basse", "DSI", "Confort utilisateur"),
]

ENTETES = ["Module", "Tâche", "Compartiment", "Avancement %", "Échéance", "Priorité", "Responsable", "Notes"]
COULEURS = {"Livré": VERTF, "En cours": AMBREF, "Planifié": GRISF}


def total_avancement():
    return round(sum(t[3] for t in TACHES) / len(TACHES))


# ------------------------------------------------------------------- Excel ---
def excel():
    wb = Workbook()
    # --- feuille de suivi ---
    ws = wb.active
    ws.title = "Suivi"
    titre = Font(bold=True, size=14, color=VIOLET)
    entete_font = Font(bold=True, color="FFFFFF")
    entete_fill = PatternFill("solid", fgColor=VIOLET)
    bord = Border(*[Side(style="thin", color="CBD5E1")] * 4)

    ws["A1"] = "VOOMNET — Avancement du projet de gestion des achats"
    ws["A1"].font = titre
    ws["A2"] = "Avancement global : %d %%   ·   %d tâches   ·   %d livrées   ·   %d en cours   ·   %d planifiées" % (
        total_avancement(), len(TACHES),
        sum(1 for t in TACHES if t[2] == "Livré"),
        sum(1 for t in TACHES if t[2] == "En cours"),
        sum(1 for t in TACHES if t[2] == "Planifié"),
    )
    ws["A2"].font = Font(italic=True, size=10, color=GRISF)

    for c, h in enumerate(ENTETES, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = entete_font
        cell.fill = entete_fill
        cell.border = bord

    for i, t in enumerate(TACHES, start=5):
        for c, v in enumerate(t, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = bord
            if c == 3:
                cell.font = Font(bold=True, color=COULEURS[t[2]])
            if c == 4:
                cell.alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4).font = Font(bold=True)

    largeurs = [16, 58, 15, 13, 12, 10, 13, 46]
    for c, l in enumerate(largeurs, 1):
        ws.column_dimensions[ws.cell(row=4, column=c).column_letter].width = l
    ws.freeze_panes = "A5"

    # --- feuilles « à coller » : une colonne par compartiment ---
    for bucket in ("Livré", "En cours", "Planifié"):
        feuille = wb.create_sheet("Coller " + bucket)
        feuille["A1"] = "À coller dans le compartiment « %s »" % bucket
        feuille["A1"].font = Font(bold=True, color=COULEURS[bucket])
        feuille.column_dimensions["A"].width = 60
        feuille.column_dimensions["B"].width = 13
        feuille.column_dimensions["C"].width = 12
        ligne = 3
        for t in TACHES:
            if t[2] != bucket:
                continue
            feuille.cell(row=ligne, column=1, value="[%s] %s (%d%%)" % (t[0], t[1], t[3]))
            feuille.cell(row=ligne, column=2, value=t[4])
            feuille.cell(row=ligne, column=3, value=t[5])
            ligne += 1
    # --- feuille technique : vrai tableau Excel nommé (requis par Power Automate) ---
    tech = wb.create_sheet("Taches")
    entetes_tech = ["Module", "Tache", "Compartiment", "Avancement", "Echeance", "Priorite", "Responsable", "Notes"]
    for c, h in enumerate(entetes_tech, 1):
        cell = tech.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=VIOLET)
    for i, t in enumerate(TACHES, start=2):
        for c, v in enumerate(t, 1):
            tech.cell(row=i, column=c, value=v)
    table = ExcelTable(displayName="TachesVOOMNET",
                       ref="A1:%s%d" % (chr(ord("A") + len(entetes_tech) - 1), len(TACHES) + 1))
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium12", showRowStripes=True)
    tech.add_table(table)
    for c, l in enumerate([16, 58, 15, 13, 12, 10, 13, 46], 1):
        tech.column_dimensions[tech.cell(row=1, column=c).column_letter].width = l

    chemin = os.path.join(SORTIE, "VOOMNET-planner.xlsx")
    wb.save(chemin)
    return chemin


# --------------------------------------------------------------------- CSV ---
def csv_file():
    chemin = os.path.join(SORTIE, "VOOMNET-planner.csv")
    with open(chemin, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(ENTETES)
        for t in TACHES:
            w.writerow(t)
    return chemin


# --------------------------------------------------------- fichiers à coller ---
def textes_a_coller():
    chemins = []
    for bucket in ("Livré", "En cours", "Planifié"):
        nom = {"Livré": "a-coller-LIVRE.txt", "En cours": "a-coller-EN-COURS.txt", "Planifié": "a-coller-PLANIFIE.txt"}[bucket]
        chemin = os.path.join(SORTIE, nom)
        with open(chemin, "w", encoding="utf-8") as f:
            f.write("À coller dans le compartiment « %s » de Microsoft Planner\n" % bucket)
            f.write("(sélectionner les lignes ci-dessous, copier, puis coller dans « Ajouter une tâche »)\n\n")
            for t in TACHES:
                if t[2] == bucket:
                    f.write("[%s] %s (%d %%)\n" % (t[0], t[1], t[3]))
        chemins.append(chemin)
    return chemins


FICHE = """# FICHE RAPIDE — Votre patron suit l'avancement dans Microsoft Planner

**Durée : 6 minutes.** Aucune installation, aucun droits d'administrateur.

## Les 6 étapes

1. **Ouvrir Planner** : https://planner.cloud.microsoft (compte professionnel Microsoft 365)

2. **Créer le plan** : `Nouveau plan` → nom : `VOOMNET — Gestion des achats` → Créer

3. **Créer 3 compartiments** : `✅ Livré` · `🔄 En cours` · `📅 Planifié`

4. **Coller les tâches livrées** :
   - cliquer dans `✅ Livré` sur **Ajouter une tâche**
   - ouvrir `a-coller-LIVRE.txt`, copier les lignes de tâches
   - coller dans le champ → **chaque ligne devient une tâche** → **Ajouter**

5. **Répéter** avec `a-coller-EN-COURS.txt` puis `a-coller-PLANIFIE.txt`

6. **Inviter votre patron** : bouton **Membres** (en haut à droite du plan) → son adresse e-mail professionnelle

## Ce qu'il verra

- **Graphiques** : répartition des tâches par statut, compartiment et personne
- **Planning** : le calendrier des échéances
- **Grille** : le détail (dates, progression, responsables)

## Avancement actuel

| Indicateur | Valeur |
|---|---|
| **Avancement global** | **{pct} %** |
| Tâches livrées / en cours / planifiées | **{liv} / {enc} / {pla}** |
| Application | déployée et en service |
| Tests | 197 contrôles au vert |

## Chaque semaine (10 min)

- passer en `✅ Livré` les tâches terminées
- mettre à jour la progression des tâches `🔄 En cours`
- ajuster les échéances

---

*Détail complet : `GUIDE-PLANNER.pdf` — 4 méthodes, dont l'automatisation Power Automate.*
"""


def fiche_rapide():
    chemin = os.path.join(SORTIE, "FICHE-RAPIDE.md")
    with open(chemin, "w", encoding="utf-8") as f:
        f.write(FICHE.format(
            pct=total_avancement(),
            liv=sum(1 for t in TACHES if t[2] == "Livré"),
            enc=sum(1 for t in TACHES if t[2] == "En cours"),
            pla=sum(1 for t in TACHES if t[2] == "Planifié")))
    return chemin


if __name__ == "__main__":
    print("Avancement global : %d %%" % total_avancement())
    for c in [excel(), csv_file(), fiche_rapide()] + textes_a_coller():
        print("✔ %s (%.1f Ko)" % (os.path.basename(c), os.path.getsize(c) / 1024))
